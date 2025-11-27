#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Fetch Facebook Page Post insights by date range (save to Excel).
- Ưu tiên dùng Page Access Token (đã truyền từ file .bat).
- Hỗ trợ day-level metrics + lifetime metrics (post_reactions_by_type_total...).
- Ghi ra .xlsx (sheet 'post_insights' + sheet 'meta').

Ví dụ chạy xem file .bat bên dưới.
"""

import os, sys, csv, time, argparse
from datetime import datetime, timedelta
import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

DEFAULT_DAY_METRICS = [
    "post_impressions",
    "post_impressions_unique",
    "post_engaged_users",
    "post_clicks",
]
DEFAULT_LIFETIME_METRICS = [
    "post_reactions_by_type_total",
]

def parse_args():
    p = argparse.ArgumentParser(description="Fetch FB Post insights and save to Excel")
    # Thông tin app/token/page
    p.add_argument("--app-id", dest="app_id", help="Facebook App ID")
    p.add_argument("--app-secret", dest="app_secret", help="Facebook App Secret")
    p.add_argument("--user-token", dest="user_token", help="User Access Token (ít dùng cho insights)")
    p.add_argument("--page-id", dest="page_id", help="Page ID (dùng để ghép post-id nếu cần)")
    p.add_argument("--page-token", dest="page_token", help="Page Access Token (nên dùng)")
    p.add_argument("--graph-version", dest="graph_version", default="v19.0", help="FB Graph API version, vd: v19.0")

    # Post & khoảng thời gian
    p.add_argument("--post-id", required=True, help="ID post (có thể là {page_id}_{post_suffix} hoặc chỉ post_suffix)")
    p.add_argument("--since", required=True, help="YYYY-MM-DD")
    p.add_argument("--until", required=True, help="YYYY-MM-DD (inclusive)")
    p.add_argument("--metrics", default=",".join(DEFAULT_DAY_METRICS + DEFAULT_LIFETIME_METRICS),
                   help="Danh sách metric, phẩy (,) ngăn cách")

    # Output
    p.add_argument("--out-xlsx", default="post_insights.xlsx", help="Đường dẫn file Excel đầu ra")
    p.add_argument("--retries", type=int, default=3, help="Số lần retry khi 429/5xx")
    return p.parse_args()

def resolve_post_id(raw_post_id, page_id):
    # Nếu người dùng truyền post suffix (không chứa '_') và có page_id thì ghép thành {page_id}_{suffix}
    if "_" in raw_post_id:
        return raw_post_id
    if page_id:
        return f"{page_id}_{raw_post_id}"
    return raw_post_id  # hy vọng là full id

def resolve_token(page_token, user_token):
    if page_token:
        return page_token, "page"
    if user_token:
        print("[WARN] Chỉ nhận được USER token. Nhiều endpoint insights yêu cầu PAGE token.", file=sys.stderr)
        return user_token, "user"
    sys.exit("Thiếu access token. Truyền --page-token (khuyến nghị) hoặc --user-token.")

def fb_get(url, params, retries=3):
    for i in range(retries):
        r = requests.get(url, params=params, timeout=30)
        # retry cho 429/5xx
        if r.status_code in (429, 500, 502, 503, 504):
            time.sleep(2 ** i)
            continue
        if r.ok:
            return r.json()
        try:
            err = r.json()
        except Exception:
            err = {"error": {"message": r.text}}
        raise RuntimeError(f"Graph error {r.status_code}: {err}")
    raise RuntimeError(f"Graph error: exceeded retries for {url}")

def daterange(since_str, until_str):
    since = datetime.strptime(since_str, "%Y-%m-%d").date()
    until = datetime.strptime(until_str, "%Y-%m-%d").date()
    if until < since:
        raise ValueError("until < since")
    d = since
    while d <= until:
        yield d
        d += timedelta(days=1)

def fetch_insights_day(post_id, metrics, token, version, since, until, retries):
    if not metrics:
        return {}
    url = f"https://graph.facebook.com/{version}/{post_id}/insights"
    params = {
        "metric": ",".join(metrics),
        "period": "day",
        "since": since,
        "until": until,
        "access_token": token,
    }
    data = fb_get(url, params, retries)
    by_date = {}
    for m in data.get("data", []):
        name = m.get("name")
        for v in m.get("values", []):
            end_time = v.get("end_time")
            date_key = end_time.split("T")[0] if end_time else None
            if not date_key:
                continue
            by_date.setdefault(date_key, {})
            by_date[date_key][name] = v.get("value")
    # đảm bảo đủ ngày
    for d in daterange(since, until):
        k = d.strftime("%Y-%m-%d")
        by_date.setdefault(k, {})
    return by_date

def fetch_insights_lifetime(post_id, metrics, token, version, retries):
    if not metrics:
        return {}
    url = f"https://graph.facebook.com/{version}/{post_id}/insights"
    params = {
        "metric": ",".join(metrics),
        "period": "lifetime",
        "access_token": token,
    }
    data = fb_get(url, params, retries)
    result = {}
    for m in data.get("data", []):
        name = m.get("name")
        vals = m.get("values", [])
        if not vals:
            continue
        v = vals[-1].get("value")
        if isinstance(v, dict):
            for k, vv in v.items():
                result[f"{name}.{k}"] = vv
        else:
            result[name] = v
    return result

def split_metrics(all_metrics):
    day, life = [], []
    for m in all_metrics:
        if m.endswith("_by_type_total") or m.endswith("_lifetime"):
            life.append(m)
        else:
            day.append(m)
    return day, life

def autosize_columns(ws):
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, length + 2), 50)

def save_to_excel(out_path, header, rows, meta: dict):
    wb = Workbook()
    ws = wb.active
    ws.title = "post_insights"
    ws.append(header)
    for r in rows:
        ws.append([r.get(c, "") for c in header])
    # Freeze & autosize
    ws.freeze_panes = "A2"
    autosize_columns(ws)

    # Sheet meta
    ws2 = wb.create_sheet("meta")
    ws2.append(["key", "value"])
    for k, v in meta.items():
        ws2.append([k, v])
    autosize_columns(ws2)

    wb.save(out_path)

def main():
    args = parse_args()

    # Chọn token
    token, token_type = resolve_token(args.page_token, args.user_token)

    # Chuẩn hóa post id
    post_id = resolve_post_id(args.post_id, args.page_id)

    all_metrics = [m.strip() for m in args.metrics.split(",") if m.strip()]
    day_metrics, lifetime_metrics = split_metrics(all_metrics)

    # Gọi API
    day_map = fetch_insights_day(
        post_id=post_id,
        metrics=day_metrics,
        token=token,
        version=args.graph_version,
        since=args.since,
        until=args.until,
        retries=args.retries
    )
    life_vals = fetch_insights_lifetime(
        post_id=post_id,
        metrics=lifetime_metrics,
        token=token,
        version=args.graph_version,
        retries=args.retries
    )

    # Chuẩn bị dữ liệu bảng
    day_cols = sorted(day_metrics)
    life_cols = sorted(life_vals.keys())
    header = ["date"] + day_cols + [f"lifetime_{c}" for c in life_cols]

    rows = []
    for date_key in sorted(day_map.keys()):
        row = {"date": date_key}
        for m in day_cols:
            row[m] = day_map[date_key].get(m, "")
        for c in life_cols:
            row[f"lifetime_{c}"] = life_vals.get(c, "")
        rows.append(row)

    # Lưu Excel
    meta = {
        "graph_version": args.graph_version,
        "token_type_used": token_type,
        "page_id": args.page_id or "",
        "post_id": post_id,
        "since": args.since,
        "until": args.until,
        "metrics": ",".join(all_metrics),
        "note": "Token ưu tiên Page token. Lifetime metrics được ghép cột với tiền tố lifetime_.",
    }
    save_to_excel(args.out_xlsx, header, rows, meta)

    print(f"[OK] Saved Excel: {args.out_xlsx}")
    print(f"Token used: {token_type}")
    print(f"Day metrics: {', '.join(day_metrics) if day_metrics else '(none)'}")
    print(f"Lifetime metrics: {', '.join(lifetime_metrics) if lifetime_metrics else '(none)'}")

if __name__ == "__main__":
    main()
